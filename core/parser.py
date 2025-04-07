import warnings
from datetime import datetime
import io

import msoffcrypto
import openpyxl
from msoffcrypto.exceptions import InvalidKeyError, DecryptionError

warnings.filterwarnings("ignore", category=UserWarning, message="Workbook contains no default style")

def parse_amount(amt):
    """Convert an Indonesian currency format (e.g., '3.246.470,00') to float."""
    if not amt or amt.strip() == '':
        return 0.0
    return float(amt.replace('.', '').replace(',', '.'))

def open_excel(filename, password):
    """Open an encrypted Excel file and return a decrypted buffer."""
    decrypted_buffer = io.BytesIO()
    try:
        with open(filename, "rb") as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_buffer)
        print("Password is correct!")
        return decrypted_buffer
    except InvalidKeyError:
        print("Incorrect password provided.")
    except DecryptionError:
        print("Decryption failed (possibly corrupted file or wrong encryption method).")
    except (OSError, ValueError) as e:
        print(f"File-related error: {e}")
    return None

def extract_transaction(row, next_row):
    """Extract a transaction dictionary from a row and its corresponding next row."""
    date_str = row[4].value.strip()
    description = row[7].value.strip() if row[7].value else ''
    incoming = row[15].value.strip() if row[15].value else ''
    outgoing = row[18].value.strip() if row[18].value else ''
    balance = row[21].value.strip() if row[21].value else ''

    time_str = next_row[4].value.split()[0] if next_row[4].value else '00:00:00'
    datetime_str = f"{date_str} {time_str}"
    datetime_obj = datetime.strptime(datetime_str, '%d %b %Y %H:%M:%S')

    return {
        'datetime': datetime_obj,
        'description': description,
        'incoming': parse_amount(incoming),
        'outgoing': parse_amount(outgoing),
        'balance': parse_amount(balance)
    }

def parse_excel_data(decrypted_buffer):
    """Parse the decrypted Excel data and extract transactions."""
    wb = openpyxl.load_workbook(decrypted_buffer)
    sheet = wb['e-Statement']

    transactions = []

    for row in sheet.iter_rows(min_row=1):
        if row[0].value is not None and isinstance(row[0].value, float):
            next_row = sheet[row[0].row + 1]
            transaction = extract_transaction(row, next_row)
            transactions.append(transaction)

    return {
        "period": sheet.cell(row=6, column=14).value,
        "transactions": transactions
    }
